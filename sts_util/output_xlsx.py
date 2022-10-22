from csv import reader
from openpyxl import load_workbook


def output(dict, char):
    print("writing xlsx file...may take few min. please wait")
    wb = load_workbook("sample.xlsx")
    ws = wb["data"]
    with open("./stsstat/merge_data.csv") as f:
        reader_l = reader(f)
        for i, row in enumerate(reader_l):
            if i == 0:
                for j, col in enumerate(row):
                    ws.cell(row=i+1, column=j+1).value = col
            else:
                if char in row:
                    for j, col in enumerate(row):
                        try:
                            ws.cell(row=i+1, column=j+1).value = float(col)
                        except:
                            ws.cell(row=i+1, column=j+1).value = col
    # del empty row
    temp = []
    for row in range(1, ws.max_row):
        if ws.cell(row=row+1, column=1).value is None:
            temp.append(row+1)
    for row in reversed(temp):
        ws.delete_rows(row)

    cell = wb["dash board"]
    cell.cell(row=3, column=3).value = char

    # pivot table, not using for loop because of diff data type
    sheet = wb["pivot table"]
    sheet.cell(row=1, column=1).value = "relic get"
    row = 2
    for i,v in dict["relic get"].items():
        sheet.cell(row=row, column=1).value = i
        sheet.cell(row=row, column=2).value = v
        row += 1

    sheet.cell(row=1, column=4).value = 'avg damage_taken'
    row = 2
    for i,v in dict['avg damage_taken'].items():
        sheet.cell(row=row, column=4).value = i
        sheet.cell(row=row, column=5).value = v
        row += 1

    sheet.cell(row=1, column=7).value = 'act 1 boss relic pick rate (%)'
    row = 2
    for i,v in dict['act 1 boss relic pick rate (%)'].items():
        sheet.cell(row=row, column=7).value = i
        sheet.cell(row=row, column=8).value = v
        row += 1

    sheet.cell(row=1, column=10).value = 'act 2 boss relic pick rate (%)'
    row = 2
    for i,v in dict['act 2 boss relic pick rate (%)'].items():
        sheet.cell(row=row, column=10).value = i
        sheet.cell(row=row, column=11).value = v
        row += 1

    sheet.cell(row=1, column=13).value = 'card pick rate act1 (exclude shop/boss/starting)'
    row = 2
    for i,v in dict['card pick rate act1 (exclude shop/boss/starting)'].items():
        sheet.cell(row=row, column=13).value = i
        sheet.cell(row=row, column=14).value = v
        row += 1

    sheet.cell(row=1, column=16).value = 'card pick rate act2 onwards (exclude shop/boss1/boss2)'
    row = 2
    for i,v in dict['card pick rate act2 onwards (exclude shop/boss1/boss2)'].items():
        sheet.cell(row=row, column=16).value = i
        sheet.cell(row=row, column=17).value = v
        row += 1

    sheet.cell(row=1, column=19).value = 'act 1 boss relic win rate (%)'
    row = 2
    for i, v in dict['act 1 boss relic win rate (%)'].items():
        sheet.cell(row=row, column=19).value = i
        sheet.cell(row=row, column=20).value = v
        row += 1

    sheet.cell(row=1, column=22).value = 'act 2 boss relic win rate (%)'
    row = 2
    for i, v in dict['act 2 boss relic win rate (%)'].items():
        sheet.cell(row=row, column=22).value = i
        sheet.cell(row=row, column=23).value = v
        row += 1

    sheet.cell(row=1, column=25).value = 'relic win rate(%)'
    row = 2
    for i, v in dict['relic win rate(%)'].items():
        sheet.cell(row=row, column=25).value = i
        sheet.cell(row=row, column=26).value = v
        row += 1

    sheet.cell(row=1, column=28).value = 'card removed'
    row = 2
    for i, v in dict['card removed'].items():
        sheet.cell(row=row, column=28).value = i
        sheet.cell(row=row, column=29).value = v
        row += 1

    sheet.cell(row=1, column=31).value = 'card win rate (exclude duplicate)'
    row = 2
    for i, v in dict['card win rate (exclude duplicate)'].items():
        sheet.cell(row=row, column=31).value = i
        sheet.cell(row=row, column=32).value = v
        row += 1

    sheet.cell(row=1, column=34).value = "card count (exclude duplicate)"
    row = 2
    for i, v in dict["card count (exclude duplicate)"].items():
        sheet.cell(row=row, column=34).value = i
        sheet.cell(row=row, column=35).value = v
        row += 1

    sheet.cell(row=1, column=37).value = 'max_damage_taken'
    row = 2
    for i in dict['max_damage_taken']:
        sheet.cell(row=row, column=37).value = i[0]
        sheet.cell(row=row, column=38).value = i[1]
        row += 1

    sheet.cell(row=1, column=40).value = 'event distribution'
    row = 2
    for i, v in dict['event distribution'].items():
        sheet.cell(row=row, column=40).value = i
        sheet.cell(row=row, column=41).value = v
        row += 1

    wb.save(f"{char}_file.xlsx")